import openpyxl
import re

def prob1And2():
    #from
    filePath = "C:\\Users\\edhnt\\Documents\\Weber Shankwick\\WSW data challenge\\diet coke.xlsx"
    workBook = openpyxl.load_workbook(filePath, read_only=True)
    sheet = workBook.active

    #to
    fileName = "C:\\Users\\edhnt\\Documents\\Weber Shankwick\\WSW data challenge\\prob1Output.xlsx"
    prob1Output = openpyxl.Workbook()
    prob1Workbook = prob1Output["Sheet"]
    outputSheet = prob1Output.active

    prob1Workbook["A1"] = "Social"
    prob1Workbook["B1"] = "News"
    prob1Workbook["C1"] = "Blog & Forum"

    addRow = 2
    socialSentiment = 0
    newsSentiment = 0
    BlogForumSentiment = 0
    socialPostives = 0
    newsPostives = 0
    BlogForumPostives = 0

    dietCokeOfficialPostColumn = 6
    lacroixOfficialPostColumn = 6
    vitaCocoOfficialPostColumn = 6
    warbyParkerCocoOfficialPostColumn = 6
    EverlaneOfficialPostColumn = 6
    venmoOfficialPostColumn = 6
    glossierOfficialPostColumn = 6
    casperOfficialPostColumn = 6
    dollarShaveClubOfficialPostColumn = 6
    birchboxOfficialPostColumn = 6
    bonobosOfficialPostColumn = 6

    cokeMentionDict = {"01":0, "02":0, "03":0, "04":0, "05":0, "06":0, "07":0, "08":0,
    "09":0, "10":0, "11":0, "12":0}
    laCroixMentionDict = {"01":0, "02":0, "03":0, "04":0, "05":0, "06":0, "07":0, "08":0,
    "09":0, "10":0, "11":0, "12":0}
    vitaCocoMentionDict = {"01":0, "02":0, "03":0, "04":0, "05":0, "06":0, "07":0, "08":0,
    "09":0, "10":0, "11":0, "12":0}
    warbyParkerMentionDict = {"01":0, "02":0, "03":0, "04":0, "05":0, "06":0, "07":0, "08":0,
    "09":0, "10":0, "11":0, "12":0}
    everlaneMentionDict = {"01":0, "02":0, "03":0, "04":0, "05":0, "06":0, "07":0, "08":0,
    "09":0, "10":0, "11":0, "12":0}
    venmoMentionDict = {"01":0, "02":0, "03":0, "04":0, "05":0, "06":0, "07":0, "08":0,
    "09":0, "10":0, "11":0, "12":0}
    glossierMentionDict = {"01":0, "02":0, "03":0, "04":0, "05":0, "06":0, "07":0, "08":0,
    "09":0, "10":0, "11":0, "12":0}
    casperMentionDict = {"01":0, "02":0, "03":0, "04":0, "05":0, "06":0, "07":0, "08":0,
    "09":0, "10":0, "11":0, "12":0}
    dollarShaveClubMentionDict = {"01":0, "02":0, "03":0, "04":0, "05":0, "06":0, "07":0, "08":0,
    "09":0, "10":0, "11":0, "12":0}
    birchboxMentionDict = {"01":0, "02":0, "03":0, "04":0, "05":0, "06":0, "07":0, "08":0,
    "09":0, "10":0, "11":0, "12":0}
    bonobosMentionDict = {"01":0, "02":0, "03":0, "04":0, "05":0, "06":0, "07":0, "08":0,
    "09":0, "10":0, "11":0, "12":0}

    cokeEngageDict = {"01":0, "02":0, "03":0, "04":0, "05":0, "06":0, "07":0, "08":0,
    "09":0, "10":0, "11":0, "12":0}
    laCroixEngageDict = {"01":0, "02":0, "03":0, "04":0, "05":0, "06":0, "07":0, "08":0,
    "09":0, "10":0, "11":0, "12":0}
    vitaCocoEngagDict = {"01":0, "02":0, "03":0, "04":0, "05":0, "06":0, "07":0, "08":0,
    "09":0, "10":0, "11":0, "12":0}
    warbyParkerEngageDict = {"01":0, "02":0, "03":0, "04":0, "05":0, "06":0, "07":0, "08":0,
    "09":0, "10":0, "11":0, "12":0}
    everlaneEngageDict = {"01":0, "02":0, "03":0, "04":0, "05":0, "06":0, "07":0, "08":0,
    "09":0, "10":0, "11":0, "12":0}
    venmoEngageDict = {"01":0, "02":0, "03":0, "04":0, "05":0, "06":0, "07":0, "08":0,
    "09":0, "10":0, "11":0, "12":0}
    glossierEngageDict = {"01":0, "02":0, "03":0, "04":0, "05":0, "06":0, "07":0, "08":0,
    "09":0, "10":0, "11":0, "12":0}
    casperEngageDict = {"01":0, "02":0, "03":0, "04":0, "05":0, "06":0, "07":0, "08":0,
    "09":0, "10":0, "11":0, "12":0}
    dollarShaveClubEngageDict = {"01":0, "02":0, "03":0, "04":0, "05":0, "06":0, "07":0, "08":0,
    "09":0, "10":0, "11":0, "12":0}
    birchboxEngageDict = {"01":0, "02":0, "03":0, "04":0, "05":0, "06":0, "07":0, "08":0,
    "09":0, "10":0, "11":0, "12":0}
    bonobosEngageDict = {"01":0, "02":0, "03":0, "04":0, "05":0, "06":0, "07":0, "08":0,
    "09":0, "10":0, "11":0, "12":0}

    for x in range(8,1000):
        #prob1
        text = sheet.cell(row=x, column=10).value
        if(text == "twitter" or text == "instagram" or text == "facebook"):
            cell = outputSheet.cell(row=addRow, column=1)
            cell.value = sheet.cell(row=x, column=6).value
            sentimentIncrease = float(sheet.cell(row=x, column=11).value)
            socialSentiment += sentimentIncrease
            if(sentimentIncrease == 1.0):
                socialPostives += 1
        elif(text == "news"):
            cell = outputSheet.cell(row=addRow, column=2)
            cell.value = sheet.cell(row=x, column=6).value
            sentimentIncrease = float(sheet.cell(row=x, column=11).value)
            newsSentiment += sentimentIncrease
            if(sentimentIncrease == 1.0):
                newsPostives += 1
        elif(text == "blog" or text == "forum"):
            cell = outputSheet.cell(row=addRow, column=3)
            cell.value = sheet.cell(row=x, column=6).value
            sentimentIncrease = float(sheet.cell(row=x, column=11).value)
            BlogForumSentiment += sentimentIncrease
            if(sentimentIncrease == 1.0):
                BlogForumPostives += 1
        addRow += 1

        #prob2
        #official posts
        author = sheet.cell(row=x, column=1).value
        if(author == "dietcoke" or author == "DietCoke" or author == "dietcokeus"):
            cell = outputSheet.cell(row=2, column=dietCokeOfficialPostColumn)
            cell.value = sheet.cell(row=x, column=6).value
            dietCokeOfficialPostColumn += 1
        elif(author == "lacroixwater" or author == "LaCroix"):
            cell = outputSheet.cell(row=3, column=lacroixOfficialPostColumn)
            cell.value = sheet.cell(row=x, column=6).value
            lacroixOfficialPostColumn += 1
        elif(author == "vitacoco" or author == "VitaCoco" or author == "VitaCocoUS"):
            cell = outputSheet.cell(row=4, column=vitaCocoOfficialPostColumn)
            cell.value = sheet.cell(row=x, column=6).value
            vitaCocoOfficialPostColumn += 1
        elif(author == "warbyparker" or author == "WarbyParker"):
            cell = outputSheet.cell(row=5, column=warbyParkerCocoOfficialPostColumn)
            cell.value = sheet.cell(row=x, column=6).value
            warbyParkerCocoOfficialPostColumn += 1
        elif(author == "everlane" or author == "Everlane"):
            cell = outputSheet.cell(row=6, column=EverlaneOfficialPostColumn)
            cell.value = sheet.cell(row=x, column=6).value
            EverlaneOfficialPostColumn += 1
        elif(author == "venmo"):
            cell = outputSheet.cell(row=7, column=venmoOfficialPostColumn)
            cell.value = sheet.cell(row=x, column=6).value
            venmoOfficialPostColumn += 1
        elif(author == "glossier"):
            cell = outputSheet.cell(row=8, column=glossierOfficialPostColumn)
            cell.value = sheet.cell(row=x, column=6).value
            glossierOfficialPostColumn += 1
        elif(author == "casper" or author == "Casper"):
            cell = outputSheet.cell(row=9, column=casperOfficialPostColumn)
            cell.value = sheet.cell(row=x, column=6).value
            casperOfficialPostColumn += 1
        elif(author == "dollarshaveclub" or author == "DollarShaveClub"):
            cell = outputSheet.cell(row=10, column=dollarShaveClubOfficialPostColumn)
            cell.value = sheet.cell(row=x, column=6).value
            dollarShaveClubOfficialPostColumn += 1
        elif(author == "birchbox" or author == "Birchbox"):
            cell = outputSheet.cell(row=11, column=birchboxOfficialPostColumn)
            cell.value = sheet.cell(row=x, column=6).value
            birchboxOfficialPostColumn += 1
        elif(author == "bonobos" or author == "Bonobos"):
            cell = outputSheet.cell(row=12, column=bonobosOfficialPostColumn)
            cell.value = sheet.cell(row=x, column=6).value
            bonobosOfficialPostColumn += 1

        #mentions and engagement
        fullText = sheet.cell(row=x, column=6).value
        blogComments = float(sheet.cell(row=x, column=2).value)
        facebookComments = float(sheet.cell(row=x, column=4).value)
        facebookLikes = float(sheet.cell(row=x, column=5).value)
        instagramComments = float(sheet.cell(row=x, column=7).value)
        instagramLikes = float(sheet.cell(row=x, column=9).value)
        twitterReplies = float(sheet.cell(row=x, column=15).value)
        twitterRetweets = float(sheet.cell(row=x, column=7).value)

        #count mentions and engagements
        if(fullText.__contains__("@dietcoke") or fullText.__contains__("@DietCoke") or fullText.__contains__("@dietcokeus")):
            month = sheet.cell(row=x, column=3).value
            cokeMentionDict[month[5:7]] += 1
            cokeEngageDict[month[5:7]] += blogComments + facebookComments + facebookLikes + instagramComments + instagramLikes + twitterReplies + twitterRetweets
        elif(fullText.__contains__("@lacroixwater") or fullText.__contains__("@LaCroix")):
            month = sheet.cell(row=x, column=3).value
            laCroixMentionDict[month[5:7]] += 1
            laCroixEngageDict[month[5:7]] += blogComments + facebookComments + facebookLikes + instagramComments + instagramLikes + twitterReplies + twitterRetweets
        elif(fullText.__contains__("@vitacoco") or fullText.__contains__("@VitaCoco") or fullText.__contains__("VitaCocoUS")):
            month = sheet.cell(row=x, column=3).value
            vitaCocoMentionDict[month[5:7]] += 1
            venmoEngageDict[month[5:7]] += blogComments + facebookComments + facebookLikes + instagramComments + instagramLikes + twitterReplies + twitterRetweets
        elif(fullText.__contains__("@warbyparker") or fullText.__contains__("@WarbyParker")):
            month = sheet.cell(row=x, column=3).value
            warbyParkerMentionDict[month[5:7]] += 1
            warbyParkerEngageDict[month[5:7]] += blogComments + facebookComments + facebookLikes + instagramComments + instagramLikes + twitterReplies + twitterRetweets
        elif(fullText.__contains__("@venmo")):
            month = sheet.cell(row=x, column=3).value
            venmoMentionDict[month[5:7]] += 1
            venmoEngageDict[month[5:7]] += blogComments + facebookComments + facebookLikes + instagramComments + instagramLikes + twitterReplies + twitterRetweets
        elif(fullText.__contains__("@glossier")):
            month = sheet.cell(row=x, column=3).value
            glossierMentionDict[month[5:7]] += 1
            glossierEngageDict[month[5:7]] += blogComments + facebookComments + facebookLikes + instagramComments + instagramLikes + twitterReplies + twitterRetweets
        elif(fullText.__contains__("@casper") or fullText.__contains__("@Casper")):
            month = sheet.cell(row=x, column=3).value
            casperMentionDict[month[5:7]] += 1
            casperEngageDict[month[5:7]] += blogComments + facebookComments + facebookLikes + instagramComments + instagramLikes + twitterReplies + twitterRetweets
        elif(fullText.__contains__("@dollarshaveclub") or fullText.__contains__("@DollarShaveClub")):
            month = sheet.cell(row=x, column=3).value
            dollarShaveClubMentionDict[month[5:7]] += 1
            dollarShaveClubEngageDict[month[5:7]] += blogComments + facebookComments + facebookLikes + instagramComments + instagramLikes + twitterReplies + twitterRetweets
        elif(fullText.__contains__("@birchbox") or fullText.__contains__("@Birchbox")):
            month = sheet.cell(row=x, column=3).value
            birchboxMentionDict[month[5:7]] += 1
            birchboxEngageDict[month[5:7]] += blogComments + facebookComments + facebookLikes + instagramComments + instagramLikes + twitterReplies + twitterRetweets
        elif(fullText.__contains__("@bonobos") or fullText.__contains__("@Bonobos")):
            month = sheet.cell(row=x, column=3).value
            bonobosMentionDict[month[5:7]] += 1
            bonobosEngageDict[month[5:7]] += blogComments + facebookComments + facebookLikes + instagramComments + instagramLikes + twitterReplies + twitterRetweets


    prob1Workbook["A1000"] = "Total Social Sentiment: " + str(socialSentiment)
    prob1Workbook["B1000"] = "Total News Sentiment: " + str(newsSentiment)
    prob1Workbook["C1000"] = "Total Blog & Forum Sentiment: " + str(BlogForumSentiment)
    categoryDict = {socialSentiment:"Social", newsSentiment:"News", BlogForumSentiment:"Blog & Forum"}
    maxSentiment = max(BlogForumSentiment, newsSentiment, socialSentiment)
    prob1Workbook["D1002"] = "Highest sentiment: " + str(maxSentiment) + " Category: " + categoryDict.get(maxSentiment)
    prob1Workbook["A1001"] = "Total Postive Social Sentiments: " + str(socialPostives)
    prob1Workbook["B1001"] = "Total Postive News Sentiments: " + str(newsPostives)
    prob1Workbook["C1001"] = "Total Postive Blog & Forum Sentiments: " + str(BlogForumPostives)
    #print official posts
    prob1Workbook["E1"] = "Official:"
    prob1Workbook["E2"] = "Diet Coke"
    prob1Workbook["E3"] = "LaCroix"
    prob1Workbook["E4"] = "Vita Coco"
    prob1Workbook["E5"] = "Warby Parker"
    prob1Workbook["E6"] = "Everlane"
    prob1Workbook["E7"] = "Venmo"
    prob1Workbook["E8"] = "Glossier"
    prob1Workbook["E9"] = "Casper"
    prob1Workbook["E10"] = "Dollar Shave Club"
    prob1Workbook["E11"] = "Birchbox"
    prob1Workbook["E12"] = "Bonobos"
    #print mentions
    #diet coke
    prob1Workbook["E13"] = "Diet Coke"
    prob1Workbook["F13"] = "Mentions per month:"
    prob1Workbook["G13"] = "01: " + str(cokeMentionDict["01"])
    prob1Workbook["H13"] = "02: " + str(cokeMentionDict["02"])
    prob1Workbook["I13"] = "03: " + str(cokeMentionDict["03"])
    prob1Workbook["J13"] = "04: " + str(cokeMentionDict["04"])
    prob1Workbook["K13"] = "05: " + str(cokeMentionDict["05"])
    prob1Workbook["L13"] = "06: " + str(cokeMentionDict["06"])
    prob1Workbook["M13"] = "07: " + str(cokeMentionDict["07"])
    prob1Workbook["N13"] = "08: " + str(cokeMentionDict["08"])
    prob1Workbook["O13"] = "09: " + str(cokeMentionDict["09"])
    prob1Workbook["P13"] = "10: " + str(cokeMentionDict["10"])
    prob1Workbook["Q13"] = "11: " + str(cokeMentionDict["11"])
    prob1Workbook["R13"] = "12: " + str(cokeMentionDict["12"])
    #lacroix
    prob1Workbook["E14"] = "LaCroix"
    prob1Workbook["F14"] = "Mentions per month:"
    prob1Workbook["G14"] = "01: " + str(laCroixMentionDict["01"])
    prob1Workbook["H14"] = "02: " + str(laCroixMentionDict["02"])
    prob1Workbook["I14"] = "03: " + str(laCroixMentionDict["03"])
    prob1Workbook["J14"] = "04: " + str(laCroixMentionDict["04"])
    prob1Workbook["K14"] = "05: " + str(laCroixMentionDict["05"])
    prob1Workbook["L14"] = "06: " + str(laCroixMentionDict["06"])
    prob1Workbook["M14"] = "07: " + str(laCroixMentionDict["07"])
    prob1Workbook["N14"] = "08: " + str(laCroixMentionDict["08"])
    prob1Workbook["O14"] = "09: " + str(laCroixMentionDict["09"])
    prob1Workbook["P14"] = "10: " + str(laCroixMentionDict["10"])
    prob1Workbook["Q14"] = "11: " + str(laCroixMentionDict["11"])
    prob1Workbook["R14"] = "12: " + str(laCroixMentionDict["12"])
    #vita coco
    prob1Workbook["E15"] = "Vita Coco"
    prob1Workbook["F15"] = "Mentions per month:"
    prob1Workbook["G15"] = "01: " + str(vitaCocoMentionDict["01"])
    prob1Workbook["H15"] = "02: " + str(vitaCocoMentionDict["02"])
    prob1Workbook["I15"] = "03: " + str(vitaCocoMentionDict["03"])
    prob1Workbook["J15"] = "04: " + str(vitaCocoMentionDict["04"])
    prob1Workbook["K15"] = "05: " + str(vitaCocoMentionDict["05"])
    prob1Workbook["L15"] = "06: " + str(vitaCocoMentionDict["06"])
    prob1Workbook["M15"] = "07: " + str(vitaCocoMentionDict["07"])
    prob1Workbook["N15"] = "08: " + str(vitaCocoMentionDict["08"])
    prob1Workbook["O15"] = "09: " + str(vitaCocoMentionDict["09"])
    prob1Workbook["P15"] = "10: " + str(vitaCocoMentionDict["10"])
    prob1Workbook["Q15"] = "11: " + str(vitaCocoMentionDict["11"])
    prob1Workbook["R15"] = "12: " + str(vitaCocoMentionDict["12"])
    #warby parker
    prob1Workbook["E16"] = "Warby Parker"
    prob1Workbook["F16"] = "Mentions per month:"
    prob1Workbook["G16"] = "01: " + str(warbyParkerMentionDict["01"])
    prob1Workbook["H16"] = "02: " + str(warbyParkerMentionDict["02"])
    prob1Workbook["I16"] = "03: " + str(warbyParkerMentionDict["03"])
    prob1Workbook["J16"] = "04: " + str(warbyParkerMentionDict["04"])
    prob1Workbook["K16"] = "05: " + str(warbyParkerMentionDict["05"])
    prob1Workbook["L16"] = "06: " + str(warbyParkerMentionDict["06"])
    prob1Workbook["M16"] = "07: " + str(warbyParkerMentionDict["07"])
    prob1Workbook["N16"] = "08: " + str(warbyParkerMentionDict["08"])
    prob1Workbook["O16"] = "09: " + str(warbyParkerMentionDict["09"])
    prob1Workbook["P16"] = "10: " + str(warbyParkerMentionDict["10"])
    prob1Workbook["Q16"] = "11: " + str(warbyParkerMentionDict["11"])
    prob1Workbook["R16"] = "12: " + str(warbyParkerMentionDict["12"])
    #everlane
    prob1Workbook["E17"] = "Everlane"
    prob1Workbook["F17"] = "Mentions per month:"
    prob1Workbook["G17"] = "01: " + str(everlaneMentionDict["01"])
    prob1Workbook["H17"] = "02: " + str(everlaneMentionDict["02"])
    prob1Workbook["I17"] = "03: " + str(everlaneMentionDict["03"])
    prob1Workbook["J17"] = "04: " + str(everlaneMentionDict["04"])
    prob1Workbook["K17"] = "05: " + str(everlaneMentionDict["05"])
    prob1Workbook["L17"] = "06: " + str(everlaneMentionDict["06"])
    prob1Workbook["M17"] = "07: " + str(everlaneMentionDict["07"])
    prob1Workbook["N17"] = "08: " + str(everlaneMentionDict["08"])
    prob1Workbook["O17"] = "09: " + str(everlaneMentionDict["09"])
    prob1Workbook["P17"] = "10: " + str(everlaneMentionDict["10"])
    prob1Workbook["Q17"] = "11: " + str(everlaneMentionDict["11"])
    prob1Workbook["R17"] = "12: " + str(everlaneMentionDict["12"])
    #venmo
    prob1Workbook["E18"] = "Venmo"
    prob1Workbook["F18"] = "Mentions per month:"
    prob1Workbook["G18"] = "01: " + str(venmoMentionDict["01"])
    prob1Workbook["H18"] = "02: " + str(venmoMentionDict["02"])
    prob1Workbook["I18"] = "03: " + str(venmoMentionDict["03"])
    prob1Workbook["J18"] = "04: " + str(venmoMentionDict["04"])
    prob1Workbook["K18"] = "05: " + str(venmoMentionDict["05"])
    prob1Workbook["L18"] = "06: " + str(venmoMentionDict["06"])
    prob1Workbook["M18"] = "07: " + str(venmoMentionDict["07"])
    prob1Workbook["N18"] = "08: " + str(venmoMentionDict["08"])
    prob1Workbook["O18"] = "09: " + str(venmoMentionDict["09"])
    prob1Workbook["P18"] = "10: " + str(venmoMentionDict["10"])
    prob1Workbook["Q18"] = "11: " + str(venmoMentionDict["11"])
    prob1Workbook["R18"] = "12: " + str(venmoMentionDict["12"])
    #glossier
    prob1Workbook["E19"] = "Glossier"
    prob1Workbook["F19"] = "Mentions per month:"
    prob1Workbook["G19"] = "01: " + str(glossierMentionDict["01"])
    prob1Workbook["H19"] = "02: " + str(glossierMentionDict["02"])
    prob1Workbook["I19"] = "03: " + str(glossierMentionDict["03"])
    prob1Workbook["J19"] = "04: " + str(glossierMentionDict["04"])
    prob1Workbook["K19"] = "05: " + str(glossierMentionDict["05"])
    prob1Workbook["L19"] = "06: " + str(glossierMentionDict["06"])
    prob1Workbook["M19"] = "07: " + str(glossierMentionDict["07"])
    prob1Workbook["N19"] = "08: " + str(glossierMentionDict["08"])
    prob1Workbook["O19"] = "09: " + str(glossierMentionDict["09"])
    prob1Workbook["P19"] = "10: " + str(glossierMentionDict["10"])
    prob1Workbook["Q19"] = "11: " + str(glossierMentionDict["11"])
    prob1Workbook["R19"] = "12: " + str(glossierMentionDict["12"])
    #casper
    prob1Workbook["E20"] = "Casper"
    prob1Workbook["F20"] = "Mentions per month:"
    prob1Workbook["G20"] = "01: " + str(casperMentionDict["01"])
    prob1Workbook["H20"] = "02: " + str(casperMentionDict["02"])
    prob1Workbook["I20"] = "03: " + str(casperMentionDict["03"])
    prob1Workbook["J20"] = "04: " + str(casperMentionDict["04"])
    prob1Workbook["K20"] = "05: " + str(casperMentionDict["05"])
    prob1Workbook["L20"] = "06: " + str(casperMentionDict["06"])
    prob1Workbook["M20"] = "07: " + str(casperMentionDict["07"])
    prob1Workbook["N20"] = "08: " + str(casperMentionDict["08"])
    prob1Workbook["O20"] = "09: " + str(casperMentionDict["09"])
    prob1Workbook["P20"] = "10: " + str(casperMentionDict["10"])
    prob1Workbook["Q20"] = "11: " + str(casperMentionDict["11"])
    prob1Workbook["R20"] = "12: " + str(casperMentionDict["12"])
    #dollar shave club
    prob1Workbook["E21"] = "Dollar Shave Club"
    prob1Workbook["F21"] = "Mentions per month:"
    prob1Workbook["G21"] = "01: " + str(dollarShaveClubMentionDict["01"])
    prob1Workbook["H21"] = "02: " + str(dollarShaveClubMentionDict["02"])
    prob1Workbook["I21"] = "03: " + str(dollarShaveClubMentionDict["03"])
    prob1Workbook["J21"] = "04: " + str(dollarShaveClubMentionDict["04"])
    prob1Workbook["K21"] = "05: " + str(dollarShaveClubMentionDict["05"])
    prob1Workbook["L21"] = "06: " + str(dollarShaveClubMentionDict["06"])
    prob1Workbook["M21"] = "07: " + str(dollarShaveClubMentionDict["07"])
    prob1Workbook["N21"] = "08: " + str(dollarShaveClubMentionDict["08"])
    prob1Workbook["O21"] = "09: " + str(dollarShaveClubMentionDict["09"])
    prob1Workbook["P21"] = "10: " + str(dollarShaveClubMentionDict["10"])
    prob1Workbook["Q21"] = "11: " + str(dollarShaveClubMentionDict["11"])
    prob1Workbook["R21"] = "12: " + str(dollarShaveClubMentionDict["12"])
    #birchbox
    prob1Workbook["E22"] = "Birchbox"
    prob1Workbook["F22"] = "Mentions per month:"
    prob1Workbook["G22"] = "01: " + str(birchboxMentionDict["01"])
    prob1Workbook["H22"] = "02: " + str(birchboxMentionDict["02"])
    prob1Workbook["I22"] = "03: " + str(birchboxMentionDict["03"])
    prob1Workbook["J22"] = "04: " + str(birchboxMentionDict["04"])
    prob1Workbook["K22"] = "05: " + str(birchboxMentionDict["05"])
    prob1Workbook["L22"] = "06: " + str(birchboxMentionDict["06"])
    prob1Workbook["M22"] = "07: " + str(birchboxMentionDict["07"])
    prob1Workbook["N22"] = "08: " + str(birchboxMentionDict["08"])
    prob1Workbook["O22"] = "09: " + str(birchboxMentionDict["09"])
    prob1Workbook["P22"] = "10: " + str(birchboxMentionDict["10"])
    prob1Workbook["Q22"] = "11: " + str(birchboxMentionDict["11"])
    prob1Workbook["R22"] = "12: " + str(birchboxMentionDict["12"])
    #bonobos
    prob1Workbook["E23"] = "Bonobos"
    prob1Workbook["F23"] = "Mentions per month:"
    prob1Workbook["G23"] = "01: " + str(bonobosMentionDict["01"])
    prob1Workbook["H23"] = "02: " + str(bonobosMentionDict["02"])
    prob1Workbook["I23"] = "03: " + str(bonobosMentionDict["03"])
    prob1Workbook["J23"] = "04: " + str(bonobosMentionDict["04"])
    prob1Workbook["K23"] = "05: " + str(bonobosMentionDict["05"])
    prob1Workbook["L23"] = "06: " + str(bonobosMentionDict["06"])
    prob1Workbook["M23"] = "07: " + str(bonobosMentionDict["07"])
    prob1Workbook["N23"] = "08: " + str(bonobosMentionDict["08"])
    prob1Workbook["O23"] = "09: " + str(bonobosMentionDict["09"])
    prob1Workbook["P23"] = "10: " + str(bonobosMentionDict["10"])
    prob1Workbook["Q23"] = "11: " + str(bonobosMentionDict["11"])
    prob1Workbook["R23"] = "12: " + str(bonobosMentionDict["12"])
    #engagements
    #coke
    prob1Workbook["S13"] = "Engagements per month:"
    prob1Workbook["T13"] = "01: " + str(cokeEngageDict["01"])
    prob1Workbook["U13"] = "02: " + str(cokeEngageDict["02"])
    prob1Workbook["V13"] = "03: " + str(cokeEngageDict["03"])
    prob1Workbook["W13"] = "04: " + str(cokeEngageDict["04"])
    prob1Workbook["X13"] = "05: " + str(cokeEngageDict["05"])
    prob1Workbook["Y13"] = "06: " + str(cokeEngageDict["06"])
    prob1Workbook["Z13"] = "07: " + str(cokeEngageDict["07"])
    prob1Workbook["AA13"] = "08: " + str(cokeEngageDict["08"])
    prob1Workbook["AB13"] = "09: " + str(cokeEngageDict["09"])
    prob1Workbook["AC13"] = "10: " + str(cokeEngageDict["10"])
    prob1Workbook["AD13"] = "11: " + str(cokeEngageDict["11"])
    prob1Workbook["AE13"] = "12: " + str(cokeEngageDict["12"])
    #lacroix
    prob1Workbook["S14"] = "Engagements per month:"
    prob1Workbook["T14"] = "01: " + str(laCroixEngageDict["01"])
    prob1Workbook["U14"] = "02: " + str(laCroixEngageDict["02"])
    prob1Workbook["V14"] = "03: " + str(laCroixEngageDict["03"])
    prob1Workbook["W14"] = "04: " + str(laCroixEngageDict["04"])
    prob1Workbook["X14"] = "05: " + str(laCroixEngageDict["05"])
    prob1Workbook["Y14"] = "06: " + str(laCroixEngageDict["06"])
    prob1Workbook["Z14"] = "07: " + str(laCroixEngageDict["07"])
    prob1Workbook["AA14"] = "08: " + str(laCroixEngageDict["08"])
    prob1Workbook["AB14"] = "09: " + str(laCroixEngageDict["09"])
    prob1Workbook["AC14"] = "10: " + str(laCroixEngageDict["10"])
    prob1Workbook["AD14"] = "11: " + str(laCroixEngageDict["11"])
    prob1Workbook["AE14"] = "12: " + str(laCroixEngageDict["12"])
    #vita coco
    prob1Workbook["S15"] = "Engagements per month:"
    prob1Workbook["T15"] = "01: " + str(vitaCocoEngagDict["01"])
    prob1Workbook["U15"] = "02: " + str(vitaCocoEngagDict["02"])
    prob1Workbook["V15"] = "03: " + str(vitaCocoEngagDict["03"])
    prob1Workbook["W15"] = "04: " + str(vitaCocoEngagDict["04"])
    prob1Workbook["X15"] = "05: " + str(vitaCocoEngagDict["05"])
    prob1Workbook["Y15"] = "06: " + str(vitaCocoEngagDict["06"])
    prob1Workbook["Z15"] = "07: " + str(vitaCocoEngagDict["07"])
    prob1Workbook["AA15"] = "08: " + str(vitaCocoEngagDict["08"])
    prob1Workbook["AB15"] = "09: " + str(vitaCocoEngagDict["09"])
    prob1Workbook["AC15"] = "10: " + str(vitaCocoEngagDict["10"])
    prob1Workbook["AD15"] = "11: " + str(vitaCocoEngagDict["11"])
    prob1Workbook["AE15"] = "12: " + str(vitaCocoEngagDict["12"])
    #warby parker
    prob1Workbook["S16"] = "Engagements per month:"
    prob1Workbook["T16"] = "01: " + str(warbyParkerEngageDict["01"])
    prob1Workbook["U16"] = "02: " + str(warbyParkerEngageDict["02"])
    prob1Workbook["V16"] = "03: " + str(warbyParkerEngageDict["03"])
    prob1Workbook["W16"] = "04: " + str(warbyParkerEngageDict["04"])
    prob1Workbook["X16"] = "05: " + str(warbyParkerEngageDict["05"])
    prob1Workbook["Y16"] = "06: " + str(warbyParkerEngageDict["06"])
    prob1Workbook["Z16"] = "07: " + str(warbyParkerEngageDict["07"])
    prob1Workbook["AA16"] = "08: " + str(warbyParkerEngageDict["08"])
    prob1Workbook["AB16"] = "09: " + str(warbyParkerEngageDict["09"])
    prob1Workbook["AC16"] = "10: " + str(warbyParkerEngageDict["10"])
    prob1Workbook["AD16"] = "11: " + str(warbyParkerEngageDict["11"])
    prob1Workbook["AE16"] = "12: " + str(warbyParkerEngageDict["12"])
    #everlane
    prob1Workbook["S17"] = "Engagements per month:"
    prob1Workbook["T17"] = "01: " + str(everlaneEngageDict["01"])
    prob1Workbook["U17"] = "02: " + str(everlaneEngageDict["02"])
    prob1Workbook["V17"] = "03: " + str(everlaneEngageDict["03"])
    prob1Workbook["W17"] = "04: " + str(everlaneEngageDict["04"])
    prob1Workbook["X17"] = "05: " + str(everlaneEngageDict["05"])
    prob1Workbook["Y17"] = "06: " + str(everlaneEngageDict["06"])
    prob1Workbook["Z17"] = "07: " + str(everlaneEngageDict["07"])
    prob1Workbook["AA17"] = "08: " + str(everlaneEngageDict["08"])
    prob1Workbook["AB17"] = "09: " + str(everlaneEngageDict["09"])
    prob1Workbook["AC17"] = "10: " + str(everlaneEngageDict["10"])
    prob1Workbook["AD17"] = "11: " + str(everlaneEngageDict["11"])
    prob1Workbook["AE17"] = "12: " + str(everlaneEngageDict["12"])
    #venmo
    prob1Workbook["S18"] = "Engagements per month:"
    prob1Workbook["T18"] = "01: " + str(venmoEngageDict["01"])
    prob1Workbook["U18"] = "02: " + str(venmoEngageDict["02"])
    prob1Workbook["V18"] = "03: " + str(venmoEngageDict["03"])
    prob1Workbook["W18"] = "04: " + str(venmoEngageDict["04"])
    prob1Workbook["X18"] = "05: " + str(venmoEngageDict["05"])
    prob1Workbook["Y18"] = "06: " + str(venmoEngageDict["06"])
    prob1Workbook["Z18"] = "07: " + str(venmoEngageDict["07"])
    prob1Workbook["AA18"] = "08: " + str(venmoEngageDict["08"])
    prob1Workbook["AB18"] = "09: " + str(venmoEngageDict["09"])
    prob1Workbook["AC18"] = "10: " + str(venmoEngageDict["10"])
    prob1Workbook["AD18"] = "11: " + str(venmoEngageDict["11"])
    prob1Workbook["AE18"] = "12: " + str(venmoEngageDict["12"])
    #glossier
    prob1Workbook["S19"] = "Engagements per month:"
    prob1Workbook["T19"] = "01: " + str(glossierEngageDict["01"])
    prob1Workbook["U19"] = "02: " + str(glossierEngageDict["02"])
    prob1Workbook["V19"] = "03: " + str(glossierEngageDict["03"])
    prob1Workbook["W19"] = "04: " + str(glossierEngageDict["04"])
    prob1Workbook["X19"] = "05: " + str(glossierEngageDict["05"])
    prob1Workbook["Y19"] = "06: " + str(glossierEngageDict["06"])
    prob1Workbook["Z19"] = "07: " + str(glossierEngageDict["07"])
    prob1Workbook["AA19"] = "08: " + str(glossierEngageDict["08"])
    prob1Workbook["AB19"] = "09: " + str(glossierEngageDict["09"])
    prob1Workbook["AC19"] = "10: " + str(glossierEngageDict["10"])
    prob1Workbook["AD19"] = "11: " + str(glossierEngageDict["11"])
    prob1Workbook["AE19"] = "12: " + str(glossierEngageDict["12"])
    #casper
    prob1Workbook["S20"] = "Engagements per month:"
    prob1Workbook["T20"] = "01: " + str(casperEngageDict["01"])
    prob1Workbook["U20"] = "02: " + str(casperEngageDict["02"])
    prob1Workbook["V20"] = "03: " + str(casperEngageDict["03"])
    prob1Workbook["W20"] = "04: " + str(casperEngageDict["04"])
    prob1Workbook["X20"] = "05: " + str(casperEngageDict["05"])
    prob1Workbook["Y20"] = "06: " + str(casperEngageDict["06"])
    prob1Workbook["Z20"] = "07: " + str(casperEngageDict["07"])
    prob1Workbook["AA20"] = "08: " + str(casperEngageDict["08"])
    prob1Workbook["AB20"] = "09: " + str(casperEngageDict["09"])
    prob1Workbook["AC20"] = "10: " + str(casperEngageDict["10"])
    prob1Workbook["AD20"] = "11: " + str(casperEngageDict["11"])
    prob1Workbook["AE20"] = "12: " + str(casperEngageDict["12"])
    #dollar shave club
    prob1Workbook["S21"] = "Engagements per month:"
    prob1Workbook["T21"] = "01: " + str(dollarShaveClubEngageDict["01"])
    prob1Workbook["U21"] = "02: " + str(dollarShaveClubEngageDict["02"])
    prob1Workbook["V21"] = "03: " + str(dollarShaveClubEngageDict["03"])
    prob1Workbook["W21"] = "04: " + str(dollarShaveClubEngageDict["04"])
    prob1Workbook["X21"] = "05: " + str(dollarShaveClubEngageDict["05"])
    prob1Workbook["Y21"] = "06: " + str(dollarShaveClubEngageDict["06"])
    prob1Workbook["Z21"] = "07: " + str(dollarShaveClubEngageDict["07"])
    prob1Workbook["AA21"] = "08: " + str(dollarShaveClubEngageDict["08"])
    prob1Workbook["AB21"] = "09: " + str(dollarShaveClubEngageDict["09"])
    prob1Workbook["AC21"] = "10: " + str(dollarShaveClubEngageDict["10"])
    prob1Workbook["AD21"] = "11: " + str(dollarShaveClubEngageDict["11"])
    prob1Workbook["AE21"] = "12: " + str(dollarShaveClubEngageDict["12"])
    #birchbox
    prob1Workbook["S22"] = "Engagements per month:"
    prob1Workbook["T22"] = "01: " + str(birchboxEngageDict["01"])
    prob1Workbook["U22"] = "02: " + str(birchboxEngageDict["02"])
    prob1Workbook["V22"] = "03: " + str(birchboxEngageDict["03"])
    prob1Workbook["W22"] = "04: " + str(birchboxEngageDict["04"])
    prob1Workbook["X22"] = "05: " + str(birchboxEngageDict["05"])
    prob1Workbook["Y22"] = "06: " + str(birchboxEngageDict["06"])
    prob1Workbook["Z22"] = "07: " + str(birchboxEngageDict["07"])
    prob1Workbook["AA22"] = "08: " + str(birchboxEngageDict["08"])
    prob1Workbook["AB22"] = "09: " + str(birchboxEngageDict["09"])
    prob1Workbook["AC22"] = "10: " + str(birchboxEngageDict["10"])
    prob1Workbook["AD22"] = "11: " + str(birchboxEngageDict["11"])
    prob1Workbook["AE22"] = "12: " + str(birchboxEngageDict["12"])
    #bonobos
    prob1Workbook["S23"] = "Engagements per month:"
    prob1Workbook["T23"] = "01: " + str(bonobosEngageDict["01"])
    prob1Workbook["U23"] = "02: " + str(bonobosEngageDict["02"])
    prob1Workbook["V23"] = "03: " + str(bonobosEngageDict["03"])
    prob1Workbook["W23"] = "04: " + str(bonobosEngageDict["04"])
    prob1Workbook["X23"] = "05: " + str(bonobosEngageDict["05"])
    prob1Workbook["Y23"] = "06: " + str(bonobosEngageDict["06"])
    prob1Workbook["Z23"] = "07: " + str(bonobosEngageDict["07"])
    prob1Workbook["AA23"] = "08: " + str(bonobosEngageDict["08"])
    prob1Workbook["AB23"] = "09: " + str(bonobosEngageDict["09"])
    prob1Workbook["AC23"] = "10: " + str(bonobosEngageDict["10"])
    prob1Workbook["AD23"] = "11: " + str(bonobosEngageDict["11"])
    prob1Workbook["AE23"] = "12: " + str(bonobosEngageDict["12"])
    prob1Output.save(fileName)

def prob3():
    #from
    filePath = "C:\\Users\\edhnt\\Documents\\Weber Shankwick\\WSW data challenge\\prob1Output.xlsx"
    workBook = openpyxl.load_workbook(filePath, read_only=True)
    sheet = workBook.active

    #to
    fileName = "C:\\Users\\edhnt\\Documents\\Weber Shankwick\\WSW data challenge\\prob3Output.xlsx"
    prob3Output = openpyxl.Workbook()
    prob3Workbook = prob3Output["Sheet"]
    outputSheet = prob3Output.active

    #regular expressions
    excitementReg = re.compile("\w*(excit(ed|ing))+\w*")
    surpriseReg = re.compile("\w*(surprised\b|amazing\b|gosh\b|omg\b|shocked\b|impress\b|thrill\b|magic\b|surprising)+\w*")
    pleasantReg = re.compile("\w*(pleasant|joyful|happy|pleasant|amused|glad|delightful|happiness)+\w*")
    refreshedReg = re.compile("\w*(refreshed|rejuvenated|relive|revived|new\s?life|nourish|refreshing)+\w*")
    energeticReg = re.compile("\w*(energetic|lively|powerful|revived|spirited|vigorous|vibrant|uplift|energized|pumped)+\w*")
    restfulReg = re.compile("\w*(restful|peaceful|calm|relaxed|unwind|destress|meditat|sooth|comfort|enjoyed|spacious|solitude|retreat|rested)+\w*")
    thankfulReg = re.compile("\w*(thankful|thank\s?god|thanks|thank\s?you)+\w*")
    angryReg = re.compile("\w*(angry\b|infuriated\b|wrath\b|unhappy\b|horrible\b|hate\b|stupid\b|scream\b|irritat\b|dread\b|pissing\b|grumbl\b|mad\b|pisses\b|pissed)+\w*")
    annoyReg = re.compile("\w*(annoyed|pique|annoyed|uncomfort|bother|miffed|irke)+\w*")
    anxietyReg = re.compile("\w*(anxious|embarrass|anxiety|afraid|concerned|wreck|jumpy|bugged|disturbed|fretful|worried|worry|depress)+\w*")
    frustrationReg = re.compile("\w*(frustrated|upset|discouraged|fouled\s?up| hung\s?up\s?on|up\s?the\s?wall|ungratified|bugged|disturbed|fretful|worried|worry|depress)+\w*")
    powerlessReg = re.compile("\w*(powerless|weak|overwhelmed|insecure|frighten|misery|resent|lonel|dreary|fatigued)+\w*")
    stressfullReg = re.compile("\w*(stressful|pressure|fatigue|distract|cortisol|burnout|meltdown|insomnia|procrastinate|mental|headach|drows|nervous|chore|sluggish|exhaust)+\w*")
    sadReg = re.compile("\w*(sore|mood\s?|sad)+\w*")

    #write topics
    prob3Workbook["A1"] = "Excitement"
    prob3Workbook["B1"] = "Surprise"
    prob3Workbook["C1"] = "Pleasant"
    prob3Workbook["D1"] = "Refreshed"
    prob3Workbook["E1"] = "Energetic"
    prob3Workbook["F1"] = "Restful"
    prob3Workbook["G1"] = "Thankful"
    prob3Workbook["H1"] = "Angry"
    prob3Workbook["I1"] = "Annoy"
    prob3Workbook["J1"] = "Anxiety"
    prob3Workbook["K1"] = "Frustration"
    prob3Workbook["L1"] = "Powerless"
    prob3Workbook["M1"] = "Stressful"
    prob3Workbook["N1"] = "Sad"

    #filter text and write to output
    for x in range(2,1000):
        fullText = sheet.cell(row=x, column=1).value
        #print(fullText)

        try:
            excitmentTest = excitementReg.search(fullText.lower())
            surpriseTest = surpriseReg.search(fullText.lower())
            pleasantTest = pleasantReg.search(fullText.lower())
            refreshedTest = refreshedReg.search(fullText.lower())
            energeticTest = energeticReg.search(fullText.lower())
            restfulTest = restfulReg.search(fullText.lower())
            thankfulTest = thankfulReg.search(fullText.lower())
            angryTest = angryReg.search(fullText.lower())
            annoyTest = annoyReg.search(fullText.lower())
            anxietyTest = anxietyReg.search(fullText.lower())
            frustrationTest = frustrationReg.search(fullText.lower())
            powerlessTest = powerlessReg.search(fullText.lower())
            stressfullTest = stressfullReg.search(fullText.lower())
            sadTest = sadReg.search(fullText.lower())

            isexcited = False
            issurprised = False
            ispleasant = False
            isrefreshed = False
            isenergetic = False
            isrestful = False
            isthankful = False
            isangry = False
            isannoy = False
            isanxiety = False
            isfrustration = False
            ispowerless = False
            isstressful = False
            issad = False

            #check if text contains strings
            if(type(excitmentTest) == re.Match):
                isexcited = True
            if(type(surpriseTest) == re.Match):
                issurprised = True
            if(type(pleasantTest) == re.Match):
                ispleasant = True
            if(type(refreshedTest) == re.Match):
                isrefreshed = True
            if(type(energeticTest) == re.Match):
                isenergetic = True
            if(type(restfulTest) == re.Match):
                isrestful = True
            if(type(thankfulTest) == re.Match):
                isthankful = True
            if(type(angryTest) == re.Match):
                isangry = True
            if(type(annoyTest) == re.Match):
                isannoy = True
            if(type(anxietyTest) == re.Match):
                isanxiety = True
            if(type(frustrationTest) == re.Match):
                isfrustration = True
            if(type(powerlessTest) == re.Match):
                ispowerless = True
            if(type(stressfullTest) == re.Match):
                isstressful = True
            if(type(sadTest) == re.Match):
                issad = True

            #output text
            if(isexcited == True):
                rowVar = 2
                while(outputSheet.cell(row=rowVar, column=1).value not in [None, "None"]):
                    rowVar += 1
                cell = outputSheet.cell(row=rowVar, column=1)
                cell.value = fullText
            if(issurprised == True):
                rowVar = 2
                while(outputSheet.cell(row=rowVar, column=2).value not in [None, "None"]):
                    rowVar += 1
                cell = outputSheet.cell(row=rowVar, column=2)
                cell.value = fullText
            if(ispleasant == True):
                rowVar = 2
                while(outputSheet.cell(row=rowVar, column=3).value not in [None, "None"]):
                    rowVar += 1
                cell = outputSheet.cell(row=rowVar, column=3)
                cell.value = fullText
            if(isrefreshed == True):
                rowVar = 2
                while(outputSheet.cell(row=rowVar, column=4).value not in [None, "None"]):
                    rowVar += 1
                cell = outputSheet.cell(row=rowVar, column=4)
                cell.value = fullText
            if(isenergetic == True):
                 rowVar = 2
                 while(outputSheet.cell(row=rowVar, column=5).value not in [None, "None"]):
                    rowVar += 1
                 cell = outputSheet.cell(row=rowVar, column=5)
                 cell.value = fullText
            if(isrestful == True):
                 rowVar = 2
                 while(outputSheet.cell(row=rowVar, column=6).value not in [None, "None"]):
                    rowVar += 1
                 cell = outputSheet.cell(row=rowVar, column=6)
                 cell.value = fullText
            if(isthankful == True):
                 rowVar = 2
                 while(outputSheet.cell(row=rowVar, column=7).value not in [None, "None"]):
                    rowVar += 1
                 cell = outputSheet.cell(row=rowVar, column=7)
                 cell.value = fullText
            if(isangry == True):
                 rowVar = 2
                 while(outputSheet.cell(row=rowVar, column=8).value not in [None, "None"]):
                    rowVar += 1
                 cell = outputSheet.cell(row=rowVar, column=8)
                 cell.value = fullText
            if(isannoy == True):
                 rowVar = 2
                 while(outputSheet.cell(row=rowVar, column=9).value not in [None, "None"]):
                    rowVar += 1
                 cell = outputSheet.cell(row=rowVar, column=9)
                 cell.value = fullText
            if(isanxiety == True):
                 rowVar = 2
                 while(outputSheet.cell(row=rowVar, column=10).value not in [None, "None"]):
                    rowVar += 1
                 cell = outputSheet.cell(row=rowVar, column=10)
                 cell.value = fullText
            if(isfrustration == True):
                 rowVar = 2
                 while(outputSheet.cell(row=rowVar, column=11).value not in [None, "None"]):
                    rowVar += 1
                 cell = outputSheet.cell(row=rowVar, column=11)
                 cell.value = fullText
            if(ispowerless == True):
                 rowVar = 2
                 while(outputSheet.cell(row=rowVar, column=12).value not in [None, "None"]):
                    rowVar += 1
                 cell = outputSheet.cell(row=rowVar, column=12)
                 cell.value = fullText
            if(isstressful == True):
                 rowVar = 2
                 while(outputSheet.cell(row=rowVar, column=13).value not in [None, "None"]):
                    rowVar += 1
                 cell = outputSheet.cell(row=rowVar, column=13)
                 cell.value = fullText
            if(issad == True):
                 rowVar = 2
                 while(outputSheet.cell(row=rowVar, column=14).value not in [None, "None"]):
                    rowVar += 1
                 cell = outputSheet.cell(row=rowVar, column=14)
                 cell.value = fullText

        except:
            print("")
    prob3Output.save(fileName)
